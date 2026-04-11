import asyncio
import json
import math
import os
import random
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, List
import queue as queue_module

import pandas as pd
import serial.tools.list_ports
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

# Try to import pyvisa — if not available, only simulation mode works
try:
    import pyvisa
    PYVISA_AVAILABLE = True
except ImportError:
    PYVISA_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = FastAPI(title="Battery Test Service", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

REPORTS_DIR = Path("./reports")
REPORTS_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------


class ConnectRequest(BaseModel):
    port: Optional[str] = None
    baud_rate: int = 115200
    simulation: bool = False


class StartRequest(BaseModel):
    order_id: str
    date: str           # "YYYY-MM"
    resistance: float   # ohms
    ocv_time: float     # seconds
    load_time: float    # seconds
    coeff: float        # K multiplier
    retest_id: Optional[int] = None  # if set, overwrite this battery index


class BatteryRecord(BaseModel):
    id: int
    ocv: float
    ccv: float
    time: str
    is_retest: bool = False


# ---------------------------------------------------------------------------
# Global session state
# ---------------------------------------------------------------------------

session = {
    "connected": False,
    "simulation": False,
    "port": None,
    "baud_rate": 115200,
    "running": False,
    "records": [],           # List of BatteryRecord dicts
    "current_readings": [],  # [(elapsed_s, voltage), ...] for live chart
    "status_text": "Ready",
    "last_ocv": None,
    "last_ccv": None,
    "order_id": None,
    "date": None,
    "retest_id": None,
}

_lock = threading.Lock()
_sse_clients: List[queue_module.Queue] = []
_test_thread: Optional[threading.Thread] = None

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _broadcast(event_dict: dict) -> None:
    """Put an event onto every SSE client queue (thread-safe)."""
    json_str = json.dumps(event_dict)
    msg = f"data: {json_str}\n\n"
    with _lock:
        clients_snapshot = list(_sse_clients)
    for q in clients_snapshot:
        try:
            q.put_nowait(msg)
        except queue_module.Full:
            pass


def _save_excel() -> None:
    """Persist session records to an Excel (or CSV) file."""
    with _lock:
        records = list(session["records"])
        order_id = session["order_id"]
        date = session["date"]

    if not order_id or not date:
        return

    if OPENPYXL_AVAILABLE:
        filepath = REPORTS_DIR / f"{order_id}_{date}.xlsx"
        wb = openpyxl.Workbook()

        # ----- RawData_Hidden sheet -----
        ws_raw = wb.active
        ws_raw.title = "RawData_Hidden"
        ws_raw.sheet_state = "hidden"
        ws_raw.append(["ID", "OCV", "CCV", "Time"])
        for rec in records:
            ws_raw.append([rec["id"], rec["ocv"], rec["ccv"], rec["time"]])

        # ----- Report sheet -----
        ws_report = wb.create_sheet("Report")
        ws_report.cell(row=14, column=1, value="ID")
        ws_report.cell(row=14, column=2, value="OCV (V)")
        ws_report.cell(row=14, column=3, value="CCV (V)")
        ws_report.cell(row=14, column=4, value="Time")

        for i, rec in enumerate(records):
            row = 15 + i
            id_cell = ws_report.cell(row=row, column=1, value=rec["id"])
            ocv_cell = ws_report.cell(row=row, column=2, value=rec["ocv"])
            ccv_cell = ws_report.cell(row=row, column=3, value=rec["ccv"])
            time_cell = ws_report.cell(row=row, column=4, value=rec["time"])
            ocv_cell.number_format = "0.000"
            ccv_cell.number_format = "0.000"

        # ----- LineChart referencing RawData_Hidden -----
        if len(records) > 0:
            chart = LineChart()
            chart.title = "OCV & CCV Trend"
            chart.style = 10
            chart.y_axis.title = "Voltage (V)"
            chart.x_axis.title = "Battery ID"

            n_rows = len(records) + 1  # including header

            ocv_data = Reference(ws_raw, min_col=2, min_row=1, max_row=n_rows)
            ccv_data = Reference(ws_raw, min_col=3, min_row=1, max_row=n_rows)
            chart.add_data(ocv_data, titles_from_data=True)
            chart.add_data(ccv_data, titles_from_data=True)

            cats = Reference(ws_raw, min_col=1, min_row=2, max_row=n_rows)
            chart.set_categories(cats)

            ws_report.add_chart(chart, "B50")

        wb.save(str(filepath))

    else:
        # Fallback: simple CSV
        filepath = REPORTS_DIR / f"{order_id}_{date}.csv"
        lines = ["ID,OCV,CCV,Time"]
        for rec in records:
            lines.append(f"{rec['id']},{rec['ocv']},{rec['ccv']},{rec['time']}")
        filepath.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Core test loop (runs in a background thread)
# ---------------------------------------------------------------------------


def _run_test_loop(params: dict) -> None:
    """
    Background thread that drives the IT8511A+ (or simulation) through
    repeated OCV → CCV test cycles until session["running"] is False.
    """
    inst = None
    try:
        simulation = params["simulation"]

        if not simulation:
            if not PYVISA_AVAILABLE:
                raise RuntimeError("pyvisa is not installed. Use simulation mode.")
            rm = pyvisa.ResourceManager()
            port = params["port"]
            # Extract numeric portion from e.g. "COM3" → 3, "/dev/ttyUSB0" stays as-is
            port_str = port
            if port.upper().startswith("COM"):
                port_num = port[3:]
                resource_str = f"ASRL{port_num}::INSTR"
            else:
                resource_str = f"ASRL{port}::INSTR"
            inst = rm.open_resource(resource_str)
            inst.baud_rate = params["baud_rate"]
            inst.data_bits = 8
            inst.stop_bits = pyvisa.constants.StopBits.one
            inst.parity = pyvisa.constants.Parity.none
            inst.write_termination = "\n"
            inst.read_termination = "\n"
            inst.timeout = 5000

            inst.write("SYST:REM")
            inst.write("INP ON")
            inst.write("MODE CURR")
            inst.write("CURR 0")

        resistance = params["resistance"]
        ocv_time = params["ocv_time"]
        load_time = params["load_time"]
        coeff = params["coeff"]

        while session["running"]:
            with _lock:
                retest_id = session["retest_id"]
                n_records = len(session["records"])

            tid = retest_id if retest_id is not None else n_records + 1

            # ------------------------------------------------------------------
            # Wait for battery insertion (voltage > 0.5 V)
            # ------------------------------------------------------------------
            _broadcast({"type": "status", "text": f"Waiting for battery ID {tid}..."})

            if simulation:
                # Simulate waiting for insertion
                for _ in range(3):
                    if not session["running"]:
                        return
                    time.sleep(0.5)
            else:
                while session["running"]:
                    try:
                        raw = inst.query("MEAS:VOLT?")
                        v = float(raw.strip())
                        if v > 0.5:
                            break
                    except Exception:
                        pass
                    time.sleep(0.1)

            if not session["running"]:
                break

            # ------------------------------------------------------------------
            # OCV phase
            # ------------------------------------------------------------------
            _broadcast({"type": "status", "text": f"OCV phase — battery {tid}"})

            last_ocv = None
            ocv_start = time.monotonic()
            while time.monotonic() - ocv_start < ocv_time:
                if not session["running"]:
                    return
                elapsed = time.monotonic() - ocv_start

                if simulation:
                    v = round(random.uniform(3.900, 4.000), 4)
                else:
                    try:
                        raw = inst.query("MEAS:VOLT?")
                        v = float(raw.strip())
                    except Exception:
                        v = 0.0

                last_ocv = v
                with _lock:
                    session["current_readings"].append((round(elapsed, 2), v))
                    session["last_ocv"] = last_ocv

                _broadcast({"type": "reading", "elapsed": round(elapsed, 2), "voltage": v, "phase": "ocv", "battery_id": tid})
                time.sleep(0.1)

            if last_ocv is None:
                last_ocv = 0.0

            # ------------------------------------------------------------------
            # Load (CCV) phase
            # ------------------------------------------------------------------
            _broadcast({"type": "status", "text": f"CCV phase — battery {tid}"})

            if not simulation:
                inst.write("MODE RES")
                inst.write(f"RES {resistance}")
                inst.write("INP ON")

            last_ccv = None
            ccv_start = time.monotonic()
            while time.monotonic() - ccv_start < load_time:
                if not session["running"]:
                    # Turn off load before exit
                    if not simulation and inst:
                        try:
                            inst.write("INP OFF")
                            inst.write("MODE CURR")
                            inst.write("CURR 0")
                        except Exception:
                            pass
                    return
                elapsed_total = (time.monotonic() - ccv_start) + ocv_time

                if simulation:
                    v_raw = (last_ocv if last_ocv is not None else 3.9) - random.uniform(0.1, 0.2)
                else:
                    try:
                        raw = inst.query("MEAS:VOLT?")
                        v_raw = float(raw.strip())
                    except Exception:
                        v_raw = 0.0

                v_display = v_raw * coeff
                last_ccv = v_raw
                with _lock:
                    session["current_readings"].append((round(elapsed_total, 2), v_display))
                    session["last_ccv"] = last_ccv

                _broadcast({"type": "reading", "elapsed": round(elapsed_total, 2), "voltage": v_display, "phase": "ccv", "battery_id": tid})
                time.sleep(0.1)

            # Turn off load
            if not simulation and inst:
                inst.write("INP OFF")
                inst.write("MODE CURR")
                inst.write("CURR 0")

            if last_ccv is None:
                last_ccv = 0.0

            # ------------------------------------------------------------------
            # Compute final values and build record
            # ------------------------------------------------------------------
            final_ccv = last_ccv * coeff
            adjusted_ocv = (last_ocv if last_ocv is not None else 0.0) * coeff
            if final_ccv > adjusted_ocv:
                final_ccv = adjusted_ocv

            now_str = datetime.now().strftime("%H:%M:%S")
            is_retest = retest_id is not None
            if is_retest:
                now_str += " (Re)"

            record = {
                "id": tid,
                "ocv": round(last_ocv, 3),
                "ccv": round(final_ccv, 3),
                "time": now_str,
                "is_retest": is_retest,
            }

            with _lock:
                if is_retest:
                    idx = tid - 1
                    if 0 <= idx < len(session["records"]):
                        session["records"][idx] = record
                    else:
                        session["records"].append(record)
                    # Auto-stop after retest completes — don't continue to next battery
                    session["running"] = False
                else:
                    session["records"].append(record)

            _broadcast({"type": "record", "record": record})
            _save_excel()

            _broadcast({"type": "status", "text": "Done - Remove battery"})

            # ------------------------------------------------------------------
            # Wait for battery removal (voltage < 0.1 V)
            # ------------------------------------------------------------------
            if simulation:
                time.sleep(2)
            else:
                while session["running"]:
                    try:
                        raw = inst.query("MEAS:VOLT?")
                        v = float(raw.strip())
                        if v < 0.1:
                            break
                    except Exception:
                        pass
                    time.sleep(0.1)

            # Reset for next cycle
            with _lock:
                session["current_readings"] = []
                session["retest_id"] = None

    except Exception as exc:
        _broadcast({"type": "error", "message": str(exc)})
        with _lock:
            session["running"] = False
    finally:
        if inst is not None:
            try:
                inst.close()
            except Exception:
                pass
        with _lock:
            session["running"] = False
        _broadcast({"type": "status", "text": "Stopped"})


# ---------------------------------------------------------------------------
# API endpoints
# ---------------------------------------------------------------------------


@app.get("/ports")
def list_ports():
    """Return a list of available serial port names."""
    ports = [p.device for p in serial.tools.list_ports.comports()]
    return {"ports": ports}


@app.post("/connect")
def connect(req: ConnectRequest):
    """Connect to the IT8511A+ (real or simulation)."""
    if req.simulation:
        with _lock:
            session["connected"] = True
            session["simulation"] = True
            session["port"] = "SIM"
            session["baud_rate"] = req.baud_rate
            session["running"] = False
        return {"ok": True, "message": "Simulation mode active"}

    if not PYVISA_AVAILABLE:
        raise HTTPException(status_code=400, detail="pyvisa is not installed. Use simulation=true.")

    if not req.port:
        raise HTTPException(status_code=400, detail="port is required when not in simulation mode.")

    try:
        rm = pyvisa.ResourceManager()
        port = req.port
        if port.upper().startswith("COM"):
            port_num = port[3:]
            resource_str = f"ASRL{port_num}::INSTR"
        else:
            resource_str = f"ASRL{port}::INSTR"

        inst = rm.open_resource(resource_str)
        inst.baud_rate = req.baud_rate
        inst.data_bits = 8
        inst.stop_bits = pyvisa.constants.StopBits.one
        inst.parity = pyvisa.constants.Parity.none
        inst.write_termination = "\n"
        inst.read_termination = "\n"
        inst.timeout = 5000
        idn = inst.query("*IDN?").strip()
        inst.close()
        rm.close()

        with _lock:
            session["connected"] = True
            session["simulation"] = False
            session["port"] = req.port
            session["baud_rate"] = req.baud_rate
            session["running"] = False

        return {"ok": True, "message": idn}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/disconnect")
def disconnect():
    """Disconnect and optionally stop any running test."""
    global _test_thread
    with _lock:
        session["running"] = False
        session["connected"] = False
        session["simulation"] = False
        session["port"] = None

    if _test_thread and _test_thread.is_alive():
        _test_thread.join(timeout=3)

    return {"ok": True}


@app.get("/status")
def get_status():
    """Return the full session state (safe copy, no queue objects)."""
    with _lock:
        state = {
            "connected": session["connected"],
            "simulation": session["simulation"],
            "port": session["port"],
            "baud_rate": session["baud_rate"],
            "running": session["running"],
            "records": list(session["records"]),
            "current_readings": list(session["current_readings"]),
            "status_text": session["status_text"],
            "last_ocv": session["last_ocv"],
            "last_ccv": session["last_ccv"],
            "order_id": session["order_id"],
            "date": session["date"],
            "retest_id": session["retest_id"],
        }
    return state


@app.post("/start")
def start_test(req: StartRequest):
    """Start the automated battery test loop."""
    global _test_thread

    if not session["connected"]:
        raise HTTPException(status_code=400, detail="Not connected")
    if session["running"]:
        raise HTTPException(status_code=400, detail="Already running")

    with _lock:
        session["order_id"] = req.order_id
        session["date"] = req.date
        session["running"] = True
        session["retest_id"] = req.retest_id
        session["status_text"] = "Running"

    params = {
        "simulation": session["simulation"],
        "port": session["port"],
        "baud_rate": session["baud_rate"],
        "resistance": req.resistance,
        "ocv_time": req.ocv_time,
        "load_time": req.load_time,
        "coeff": req.coeff,
    }

    _test_thread = threading.Thread(target=_run_test_loop, args=(params,), daemon=True)
    _test_thread.start()

    return {"ok": True}


@app.post("/stop")
def stop_test():
    """Signal the test loop to stop."""
    with _lock:
        session["running"] = False
    return {"ok": True}


@app.get("/stream")
def stream_events():
    """
    Server-Sent Events endpoint.

    Uses a threading.Queue so the background thread can push events without
    needing to interact with the asyncio event loop.
    """
    client_queue: queue_module.Queue = queue_module.Queue(maxsize=256)

    with _lock:
        _sse_clients.append(client_queue)

    def event_generator():
        heartbeat_interval = 15  # seconds
        last_heartbeat = time.monotonic()
        try:
            while True:
                try:
                    msg = client_queue.get(timeout=1)
                    yield msg
                    last_heartbeat = time.monotonic()
                except queue_module.Empty:
                    # Send heartbeat if idle for too long
                    if time.monotonic() - last_heartbeat >= heartbeat_interval:
                        yield ": heartbeat\n\n"
                        last_heartbeat = time.monotonic()
        except GeneratorExit:
            pass
        finally:
            with _lock:
                if client_queue in _sse_clients:
                    _sse_clients.remove(client_queue)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/report/download")
def download_report():
    """Download the Excel (or CSV) report for the current session."""
    with _lock:
        order_id = session["order_id"]
        date = session["date"]

    if not order_id or not date:
        raise HTTPException(status_code=404, detail="No report available — start a test first.")

    xlsx_path = REPORTS_DIR / f"{order_id}_{date}.xlsx"
    csv_path = REPORTS_DIR / f"{order_id}_{date}.csv"

    if xlsx_path.exists():
        return FileResponse(
            path=str(xlsx_path),
            filename=xlsx_path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    if csv_path.exists():
        return FileResponse(
            path=str(csv_path),
            filename=csv_path.name,
            media_type="text/csv",
        )

    raise HTTPException(status_code=404, detail="Report file not found.")


@app.delete("/session")
def clear_session():
    """Stop any running test and wipe all session data."""
    global _test_thread

    with _lock:
        session["running"] = False

    if _test_thread and _test_thread.is_alive():
        _test_thread.join(timeout=3)

    with _lock:
        session["records"] = []
        session["current_readings"] = []
        session["status_text"] = "Ready"
        session["last_ocv"] = None
        session["last_ccv"] = None
        session["order_id"] = None
        session["date"] = None
        session["retest_id"] = None

    return {"ok": True}
